from openpyxl import Workbook
import jsonpickle
import random
import string
import conftest
import os
import getopt
import sys

from model.user import User

try:
    opts, args = getopt.getopt(sys.argv[1:], "n:f", ["number of users", "file"])
except getopt.GetoptError as err:
    getopt.usage()
    sys.exit(2)

n = 5
f = "data/usersExcel.xlsx"

for o, a in opts:
    if o == "-n":
        n = int(a)
    elif o == "-f":
        f = a
        if f is "":
            f = str(args[0])

def random_string(prefix, maxlen):
    symbols = string.ascii_letters + string.digits + " " * 10
    return prefix + "".join([random.choice(symbols) for i in range(random.randrange(maxlen))])


testData = [
    User(name=random_string("name", 10), last_name=random_string("last", 10), nick_name=random_string("nick", 10))
    for i in range(n)
]

file = os.path.join(conftest.ROOT_DIR, f)

# with open(file, "w") as out:
#     jsonpickle.set_encoder_options("json", indent=2)
#     out.write(jsonpickle.encode(testData))

# with open(file, "w") as out:
wb = Workbook()
wSheet1 = wb.active
wSheet1.title = "data"

for row in range(1, len(testData) + 1):
    _=wSheet1.cell(column=1, row=row, value=str(testData[row-1].name))
    _=wSheet1.cell(column=2, row=row, value=str(testData[row-1].last_name))
    _=wSheet1.cell(column=3, row=row, value=str(testData[row-1].nick_name))

wb.save(f)
